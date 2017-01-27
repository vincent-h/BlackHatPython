import netaddr
import os
import openpyxl
import socket

os.chdir("/Users/Vincent1/Desktop")
# print netaddr.IPAddress()
ip_list = []
dns_list = []
ip = netaddr.IPNetwork(raw_input("Geef ip adres en subnetmask op (ip/subnet): "))

for ip in list(ip):
  ip_addr = str(ip)
  response = os.system("ping -c1 -W0.1 " + ip_addr)

  if response == 0:
    ip_list.append(str(ip))
    hostname = socket.gethostbyaddr(str(ip))[0]
    dns_list.append(hostname)
  else:
    pass

# Write to excel file
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = "Active IP Adresses"
ws['C1'] = "Hostname"
ws.append([])

#adds the ip and hostname to the excel file before saving.
for ip in ip_list:
    counter = 0 # as ip_list is 1:1 length with dns_list, this was the easiest workaround.
    ws.append([ip, '', dns_list[counter]])
    counter += 1
wb.save(raw_input("Geef een bestandsnaam op: ")+ ".xlsx")
